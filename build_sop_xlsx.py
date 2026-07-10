from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BLUE = Font(name='Arial', color='0000FF')          # inputs
BLACK = Font(name='Arial', color='000000')
HDR = Font(name='Arial', bold=True, color='FFFFFF')
HFILL = PatternFill('solid', start_color='2E5A2E')  # chocolate-green header
TITLE = Font(name='Arial', bold=True, size=13, color='2E5A2E')
NOTE = Font(name='Arial', italic=True, size=9, color='666666')
thin = Side(style='thin', color='CCCCCC')
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def hdr_row(ws, row, headers, start=1):
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=start+j, value=h)
        c.font = HDR; c.fill = HFILL; c.alignment = Alignment(horizontal='center'); c.border = BORD

def put(ws, row, vals, start=1, font=BLUE, numfmt=None):
    for j, v in enumerate(vals):
        c = ws.cell(row=row, column=start+j, value=v)
        c.font = font if not isinstance(v, str) else BLACK
        c.border = BORD
        if numfmt and isinstance(v, (int, float)): c.number_format = numfmt

# ---------- 1. CONFIG / SKU ----------
ws = wb.active; ws.title = '1_Config_SKU'
ws['A1'] = 'S&OP / IBP — Maestro del SKU (INPUT manual)'; ws['A1'].font = TITLE
hdr_row(ws, 3, ['Parametro', 'Valor', 'Unidad'])
cfg = [
    ('marca', 'ChocoAndes', 'texto'),
    ('sku', 'BARRA-70CACAO-100G', 'texto'),
    ('descripcion', 'Barra chocolate 70% cacao 100g', 'texto'),
    ('hemisferio', 'sur', 'texto (sur=peak invierno)'),
    ('horizonte_meses', 12, 'meses'),
    ('precio_venta_unitario', 1890, 'CLP/unidad'),
    ('moneda', 'CLP', 'texto'),
]
for i, r in enumerate(cfg): put(ws, 4+i, r)
ws['A12'] = 'Azul = input editable por el planner'; ws['A12'].font = NOTE
for col,w in {'A':26,'B':30,'C':26}.items(): ws.column_dimensions[col].width = w

# ---------- 2. DEMAND HISTORY ----------
ws = wb.create_sheet('2_Demanda_Historico')
ws['A1'] = 'Histórico de ventas (INPUT) — base estadística de la proyección'; ws['A1'].font = TITLE
hdr_row(ws, 3, ['anio', 'mes', 'mes_nombre', 'unidades_vendidas'])
nombres = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
hist = [
 (2024,1,8200),(2024,2,7900),(2024,3,11200),(2024,4,10100),(2024,5,12800),(2024,6,15600),
 (2024,7,16200),(2024,8,15100),(2024,9,12400),(2024,10,11800),(2024,11,12900),(2024,12,17800),
 (2025,1,8800),(2025,2,8400),(2025,3,12100),(2025,4,11000),(2025,5,13700),(2025,6,16800),
 (2025,7,17500),(2025,8,16300),(2025,9,13300),(2025,10,12600),(2025,11,13900),(2025,12,19100),
]
for i,(a,m,u) in enumerate(hist):
    put(ws, 4+i, [a, m, nombres[m-1], u])
    ws.cell(row=4+i, column=3).font = BLACK
ws['A30'] = 'Estacionalidad chocolate (Chile): peak invierno Jun-Ago + Pascua (Mar/Abr) + Navidad (Dic); valle verano Ene-Feb.'; ws['A30'].font = NOTE
for col,w in {'A':8,'B':6,'C':12,'D':18}.items(): ws.column_dimensions[col].width = w

# ---------- 3. CAPACIDAD FABRICA ----------
ws = wb.create_sheet('3_Capacidad_Fabrica')
ws['A1'] = 'Capacidad de fábrica (INPUT manual) — restricción de producción'; ws['A1'].font = TITLE
hdr_row(ws, 3, ['Parametro', 'Valor', 'Unidad'])
cap = [
 ('lineas_produccion', 2, 'líneas'),
 ('unidades_por_hora_linea', 850, 'u/hora/línea'),
 ('horas_turno', 8, 'horas'),
 ('turnos_por_dia_normal', 1, 'turnos'),
 ('dias_habiles_mes', 22, 'días'),
 ('capacidad_mensual_normal', 374000, 'u/mes (calc: 2*850*8*1*22)'),
 ('turno_extra_disponible', 'si', 'si/no'),
 ('capacidad_mensual_con_turno_extra', 748000, 'u/mes (2 turnos)'),
 ('inventario_inicial', 4000, 'unidades'),
 ('stock_seguridad_objetivo', 2500, 'unidades'),
]
for i,r in enumerate(cap): put(ws, 4+i, r)
for col,w in {'A':34,'B':16,'C':30}.items(): ws.column_dimensions[col].width = w

# ---------- 4. MATERIAS PRIMAS ----------
ws = wb.create_sheet('4_Materias_Primas')
ws['A1'] = 'Restricciones de materia prima e insumos (INPUT manual)'; ws['A1'].font = TITLE
hdr_row(ws, 3, ['insumo','consumo_por_unidad','unidad','disponibilidad_mensual','unidad_disp','lead_time_dias'])
mp = [
 ('Cacao 70%', 0.070, 'kg/u', 1200, 'kg/mes', 45),
 ('Leche en polvo', 0.020, 'kg/u', 600, 'kg/mes', 20),
 ('Azúcar', 0.025, 'kg/u', 900, 'kg/mes', 15),
 ('Envase/wrapper', 1.0, 'u/u', 20000, 'u/mes', 30),
 ('Caja display', 0.042, 'u/u', 1000, 'u/mes', 25),
]
for i,r in enumerate(mp):
    put(ws, 4+i, r)
    ws.cell(row=4+i,column=3).font=BLACK; ws.cell(row=4+i,column=5).font=BLACK
ws['A11'] = 'El cuello de botella define el máximo producible: min(disponibilidad / consumo_por_unidad).'; ws['A11'].font = NOTE
for col,w in {'A':18,'B':20,'C':10,'D':22,'E':12,'F':16}.items(): ws.column_dimensions[col].width = w

# ---------- 5. LOGISTICA ----------
ws = wb.create_sheet('5_Logistica')
ws['A1'] = 'Restricciones logísticas (INPUT manual)'; ws['A1'].font = TITLE
hdr_row(ws, 3, ['Parametro','Valor','Unidad'])
log = [
 ('capacidad_bodega_PT', 30000, 'u (producto terminado)'),
 ('capacidad_despacho_mensual', 22000, 'u/mes (flota)'),
 ('unidades_por_pallet', 480, 'u/pallet'),
 ('posiciones_pallet_bodega', 62, 'pallets'),
 ('lead_time_distribucion_dias', 5, 'días'),
]
for i,r in enumerate(log): put(ws, 4+i, r)
for col,w in {'A':30,'B':14,'C':26}.items(): ws.column_dimensions[col].width = w

# ---------- 6. SUPUESTOS FINANCIEROS ----------
ws = wb.create_sheet('6_Supuestos_Financieros')
ws['A1'] = 'Supuestos financieros (INPUT manual) — para el análisis P&L'; ws['A1'].font = TITLE
hdr_row(ws, 3, ['Parametro','Valor','Unidad'])
fin = [
 ('precio_venta_unitario', 1890, 'CLP/u'),
 ('costo_materia_prima_unitario', 620, 'CLP/u'),
 ('costo_mano_obra_unitario', 240, 'CLP/u'),
 ('costo_overhead_unitario', 240, 'CLP/u'),
 ('costo_total_unitario (COGS)', 1100, 'CLP/u'),
 ('costo_logistica_unitario', 95, 'CLP/u'),
 ('gasto_marketing_mensual', 3500000, 'CLP/mes'),
 ('costo_quiebre_oportunidad', 'precio*unid_no_servidas', 'regla'),
 ('tasa_descuento_anual', 0.12, '% (para VAN si se requiere)'),
]
for i,r in enumerate(fin): put(ws, 4+i, r)
for col,w in {'A':34,'B':24,'C':24}.items(): ws.column_dimensions[col].width = w

wb.save('SOP_Chocolate_INPUTS.xlsx')
print('saved')
